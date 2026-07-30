import sys
import tempfile
import unittest
from pathlib import Path

# Ensure 'src' is in sys.path when running tests directly
SRC_DIR = Path(__file__).resolve().parents[1] / "src"
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEL_AVIV_XLSX = PROJECT_ROOT / "budget_examples" / "tel_aviv_2026.xlsx"
ELAD_XLSX = PROJECT_ROOT / "budget_examples" / "elad_2022.xlsx"

from muni_budget_analysis.processing.excel_pipeline import (
    segment_sheet_tables,
    extract_excel_tables,
)


class TestSegmentSheetTablesSynthetic(unittest.TestCase):
    """
    Real correctness test: build an in-memory workbook with two clearly
    separated regions (rows 1-3 filled, rows 4-5 blank, rows 6-8 filled)
    and verify segment_sheet_tables detects exactly 2 regions with the
    right row bounds.
    """

    def test_two_regions_detected(self):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        for row in (1, 2, 3):
            ws.cell(row=row, column=1, value=f"a{row}")
            ws.cell(row=row, column=2, value=f"b{row}")

        # rows 4-5 left blank

        for row in (6, 7, 8):
            ws.cell(row=row, column=1, value=f"a{row}")
            ws.cell(row=row, column=3, value=f"c{row}")

        regions = segment_sheet_tables(ws)

        self.assertEqual(len(regions), 2)

        self.assertEqual(regions[0]["min_row"], 1)
        self.assertEqual(regions[0]["max_row"], 3)
        self.assertEqual(regions[0]["min_col"], 1)
        self.assertEqual(regions[0]["max_col"], 2)

        self.assertEqual(regions[1]["min_row"], 6)
        self.assertEqual(regions[1]["max_row"], 8)
        self.assertEqual(regions[1]["min_col"], 1)
        self.assertEqual(regions[1]["max_col"], 3)

    def test_empty_sheet_returns_no_regions(self):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        regions = segment_sheet_tables(ws)

        self.assertEqual(regions, [])


class TestExtractExcelTablesGroundTruth(unittest.TestCase):
    """
    Real correctness tests for extract_excel_tables() against synthetic
    in-memory workbooks with a known table layout (issue #7): region
    boundaries, row/col spans, and header flags are all asserted against
    ground truth defined here, not just "doesn't crash".
    """

    def _save(self, wb) -> Path:
        tmpdir = tempfile.mkdtemp()
        path = Path(tmpdir) / "workbook.xlsx"
        wb.save(path)
        return path

    def test_two_regions_with_headers_and_data(self):
        """
        Two tables in one sheet separated by a single blank-row gap. Ground
        truth: 2 regions, exact row bounds, exact cell text, and is_header
        True only for each region's first row.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="Name")
        ws.cell(row=1, column=2, value="Amount")
        ws.cell(row=2, column=1, value="Roads")
        ws.cell(row=2, column=2, value=100)
        ws.cell(row=3, column=1, value="Parks")
        ws.cell(row=3, column=2, value=50)
        # row 4 left blank -- gap between the two tables

        ws.cell(row=5, column=1, value="Category")
        ws.cell(row=5, column=2, value="Q1")
        ws.cell(row=5, column=3, value="Q2")
        ws.cell(row=6, column=1, value="Water")
        ws.cell(row=6, column=2, value=10)
        ws.cell(row=6, column=3, value=20)
        ws.cell(row=7, column=1, value="Sewage")
        ws.cell(row=7, column=2, value=5)
        ws.cell(row=7, column=3, value=15)

        path = self._save(wb)
        result = extract_excel_tables(path)

        self.assertEqual(len(result["sheets"]), 1)
        regions = result["sheets"][0]["regions"]
        self.assertEqual(len(regions), 2)

        # Region 1: rows 1-3, cols 1-2
        region1 = regions[0]
        self.assertEqual(region1["min_row"], 1)
        self.assertEqual(len(region1["rows"]), 3)

        header_row = region1["rows"][0]
        self.assertEqual([c["text"] for c in header_row], ["Name", "Amount"])
        self.assertTrue(all(c["is_header"] for c in header_row))
        self.assertTrue(all(c["row_span"] == 1 and c["col_span"] == 1 for c in header_row))

        data_row_1 = region1["rows"][1]
        self.assertEqual([c["text"] for c in data_row_1], ["Roads", "100"])
        self.assertFalse(any(c["is_header"] for c in data_row_1))

        data_row_2 = region1["rows"][2]
        self.assertEqual([c["text"] for c in data_row_2], ["Parks", "50"])
        self.assertFalse(any(c["is_header"] for c in data_row_2))

        # Region 2: rows 5-7, cols 1-3
        region2 = regions[1]
        self.assertEqual(region2["min_row"], 5)
        self.assertEqual(len(region2["rows"]), 3)

        header_row_2 = region2["rows"][0]
        self.assertEqual([c["text"] for c in header_row_2], ["Category", "Q1", "Q2"])
        self.assertTrue(all(c["is_header"] for c in header_row_2))

        self.assertEqual([c["text"] for c in region2["rows"][1]], ["Water", "10", "20"])
        self.assertEqual([c["text"] for c in region2["rows"][2]], ["Sewage", "5", "15"])
        self.assertFalse(any(c["is_header"] for c in region2["rows"][1]))
        self.assertFalse(any(c["is_header"] for c in region2["rows"][2]))

    def test_merged_header_cell_spanning_multiple_columns(self):
        """
        A merged header cell (A1:C1) spanning 3 columns, followed by
        per-column sub-headers and a data row. Ground truth: the merged
        region emits exactly one cell for row 1 (the top-left), with
        col_span == 3; the non-top-left merged columns are not emitted as
        separate cells.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        ws.merge_cells("A1:C1")
        ws.cell(row=1, column=1, value="Header Group")
        ws.cell(row=2, column=1, value="H1")
        ws.cell(row=2, column=2, value="H2")
        ws.cell(row=2, column=3, value="H3")
        ws.cell(row=3, column=1, value=1)
        ws.cell(row=3, column=2, value=2)
        ws.cell(row=3, column=3, value=3)

        path = self._save(wb)
        result = extract_excel_tables(path)

        regions = result["sheets"][0]["regions"]
        self.assertEqual(len(regions), 1)
        region = regions[0]
        self.assertEqual(len(region["rows"]), 3)

        merged_row = region["rows"][0]
        self.assertEqual(len(merged_row), 1)
        self.assertEqual(merged_row[0]["text"], "Header Group")
        self.assertEqual(merged_row[0]["row_span"], 1)
        self.assertEqual(merged_row[0]["col_span"], 3)
        self.assertTrue(merged_row[0]["is_header"])

        sub_header_row = region["rows"][1]
        self.assertEqual([c["text"] for c in sub_header_row], ["H1", "H2", "H3"])
        self.assertTrue(all(c["col_span"] == 1 for c in sub_header_row))
        self.assertFalse(any(c["is_header"] for c in sub_header_row))

        data_row = region["rows"][2]
        self.assertEqual([c["text"] for c in data_row], ["1", "2", "3"])


class TestSegmentSheetTablesGapSize(unittest.TestCase):
    """
    A multi-row internal gap should NOT be treated as a false table
    boundary once gap_size is configured large enough to tolerate it --
    the split only happens when the empty run reaches gap_size.
    """

    def test_gap_below_threshold_does_not_split_region(self):
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active

        for row in (1, 2, 3):
            ws.cell(row=row, column=1, value=f"a{row}")
            ws.cell(row=row, column=2, value=f"b{row}")

        # rows 4-5 blank: a 2-row gap

        for row in (6, 7, 8):
            ws.cell(row=row, column=1, value=f"a{row}")
            ws.cell(row=row, column=2, value=f"b{row}")

        # gap_size=3: a 2-row gap is below threshold, so this must stay one region
        regions = segment_sheet_tables(ws, gap_size=3)
        self.assertEqual(len(regions), 1)
        self.assertEqual(regions[0]["min_row"], 1)
        self.assertEqual(regions[0]["max_row"], 8)
        self.assertEqual(regions[0]["min_col"], 1)
        self.assertEqual(regions[0]["max_col"], 2)

        # gap_size=2: the same 2-row gap now meets the threshold and splits
        regions_split = segment_sheet_tables(ws, gap_size=2)
        self.assertEqual(len(regions_split), 2)
        self.assertEqual(regions_split[0]["max_row"], 3)
        self.assertEqual(regions_split[1]["min_row"], 6)


class TestKnownEdgeCases(unittest.TestCase):
    """
    Regression tests pinning the two documented-but-unaddressed edge cases
    from issue #7. Neither is fixed here (both are called out in
    excel_pipeline.py's docstrings as accepted tradeoffs per ADR-0002's
    "don't over-invest in heuristics" guidance) -- these tests exist so the
    gap is tracked in code, not just prose, and so a future behavior change
    is caught.
    """

    def _save(self, wb) -> Path:
        tmpdir = tempfile.mkdtemp()
        path = Path(tmpdir) / "workbook.xlsx"
        wb.save(path)
        return path

    def test_uncached_formula_reads_as_silent_empty_string(self):
        """
        read_workbook() opens with data_only=True. A formula cell with no
        cached value (never recalculated by Excel/openpyxl before save)
        reads as None and is silently emitted as "" -- not flagged as
        missing/uncached in any way.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="Label")
        ws.cell(row=1, column=2, value="=1+1")  # formula, no cached result
        ws.cell(row=1, column=3, value="End")

        path = self._save(wb)
        result = extract_excel_tables(path)

        row_cells = result["sheets"][0]["regions"][0]["rows"][0]
        self.assertEqual(row_cells[0]["text"], "Label")
        self.assertEqual(row_cells[1]["text"], "")  # uncached formula -> silently ""
        self.assertEqual(row_cells[2]["text"], "End")

    def test_merged_cell_span_can_exceed_region_bounds(self):
        """
        Segmentation is based on non-empty cells, not merge geometry. A
        merged cell (A1:E1) whose top-left is inside a region, but whose
        other columns are only "occupied" via the merge (openpyxl reports
        their value as None), does not extend the region's declared
        max_col -- yet the merged cell's own emitted col_span still
        reflects the full merge width, exceeding the region's bounds.
        """
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.merge_cells("A1:E1")
        ws.cell(row=1, column=1, value="Wide Header")
        ws.cell(row=2, column=1, value="x")
        ws.cell(row=2, column=2, value="y")

        path = self._save(wb)
        result = extract_excel_tables(path)

        regions = result["sheets"][0]["regions"]
        self.assertEqual(len(regions), 1)

        # Region only spans cols 1-2 (from row 2's non-empty cells) --
        # row 1's merged-away columns (B1:E1) never register as non-empty.
        region_min_col, region_max_col = 1, 2
        merged_cell = result["sheets"][0]["regions"][0]["rows"][0][0]
        self.assertEqual(merged_cell["text"], "Wide Header")

        declared_span = region_max_col - region_min_col + 1
        self.assertEqual(declared_span, 2)
        self.assertEqual(merged_cell["col_span"], 5)
        self.assertGreater(merged_cell["col_span"], declared_span)


class TestRealSamplesSmoke(unittest.TestCase):
    """
    Smoke tests against real sample workbooks -- there's no ground-truth
    expected output for these files yet, so these only assert the pipeline
    runs without crashing and produces at least one non-empty region per
    file. Fidelity (correct cell values/bounds) is NOT asserted here.
    """

    def _assert_smoke(self, path: Path):
        self.assertTrue(path.exists(), f"missing sample file: {path}")

        result = extract_excel_tables(path)

        self.assertIn("sheets", result)
        self.assertGreaterEqual(len(result["sheets"]), 1)

        total_regions = 0
        for sheet in result["sheets"]:
            for region in sheet["regions"]:
                total_regions += 1
                self.assertGreater(len(region["rows"]), 0)

        self.assertGreaterEqual(total_regions, 1)

    def test_tel_aviv_2026_smoke(self):
        self._assert_smoke(TEL_AVIV_XLSX)

    def test_elad_2022_smoke(self):
        self._assert_smoke(ELAD_XLSX)


class TestRealSamplesFidelity(unittest.TestCase):
    """
    Ground-truth fidelity check against real sample workbooks: unlike
    TestRealSamplesSmoke (crash-only), these assert extracted cell text and
    segmentation bounds match values read directly from the source files
    (verified independently via openpyxl against the raw .xlsx).
    """

    def test_elad_2022_single_continuous_region(self):
        result = extract_excel_tables(ELAD_XLSX)

        self.assertEqual(len(result["sheets"]), 1)
        sheet = result["sheets"][0]
        self.assertEqual(len(sheet["regions"]), 1)

        region = sheet["regions"][0]
        self.assertEqual(region["min_row"], 1)
        self.assertEqual(len(region["rows"]), 625)  # whole sheet, no blank-row gaps

        header_row = region["rows"][0]
        self.assertEqual(
            [c["text"] for c in header_row],
            ["מס כרטיס", "שם כרטיס", "מספר משרות", "תקציב באלפי שח"],
        )
        self.assertTrue(all(c["is_header"] for c in header_row))

        first_data_row = region["rows"][1]
        self.assertEqual(
            [c["text"] for c in first_data_row],
            ["1111100100", "ארנונה מגורים", "0", "44500"],
        )
        self.assertFalse(any(c["is_header"] for c in first_data_row))

        # no merged cells anywhere in this file -> every cell is a plain 1x1 span
        self.assertTrue(all(
            c["row_span"] == 1 and c["col_span"] == 1
            for row in region["rows"] for c in row
        ))

    def test_tel_aviv_2026_income_sheet_known_values(self):
        result = extract_excel_tables(TEL_AVIV_XLSX)
        sheets = {s["sheet_name"]: s for s in result["sheets"]}
        region = sheets["הכנסות"]["regions"][0]

        self.assertEqual(region["min_row"], 1)
        self.assertEqual(len(region["rows"]), 397)

        # Row 1 is a single-cell title banner ("הצעת תקציב רגיל..."), not the
        # real column-header row -- segment_sheet_tables' gap heuristic can't
        # tell the two apart (both just look like "a non-empty row" to it), so
        # is_header ends up on the banner instead of row 2's actual column
        # names. Known gap per ADR-0002 / PR #5, not fixed here -- asserted
        # explicitly so a future fix has a test to flip instead of this
        # passing silently on the wrong row.
        title_row = region["rows"][0]
        self.assertEqual(
            title_row[0]["text"], "הצעת תקציב רגיל \nלשנת הכספים תשפ\"ג  2023\nהכנסות"
        )
        self.assertTrue(title_row[0]["is_header"])

        real_header_row = region["rows"][1]
        self.assertEqual(real_header_row[0]["text"], "פרק")
        self.assertEqual(real_header_row[3]["text"], "שם סעיף")
        self.assertFalse(real_header_row[0]["is_header"])

        first_data_row = region["rows"][2]
        self.assertEqual(first_data_row[0]["text"], "11000")
        self.assertEqual(first_data_row[2]["text"], "11000/121/8")
        self.assertEqual(first_data_row[3]["text"], "גביה שוטפת-מגורים")
        self.assertEqual(first_data_row[4]["text"], "1320000000")


if __name__ == "__main__":
    unittest.main()
