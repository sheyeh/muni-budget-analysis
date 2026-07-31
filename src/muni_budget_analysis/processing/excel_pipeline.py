"""
excel_pipeline.py

Native openpyxl read + table-region segmentation for Excel budget files
(per docs/adr/0002-excel-native-extraction-path.md). Excel provides exact
cell values, so unlike the PDF path there's no ML layout inference here --
just a gap-based heuristic to split each sheet into table regions.

The gap-based segmentation heuristic is explicitly provisional/unvalidated
(see ADR-0002) -- it exists to give stage 3 something table-shaped to work
with, not as a general-purpose table detector.
"""

import logging
from pathlib import Path
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)


def read_workbook(path: Path):
    """
    Open an xlsx/xls workbook with formulas resolved to values.

    Not opened with read_only=True: openpyxl's read-only worksheets don't
    expose `merged_cells`, which extract_excel_tables() needs for row/col
    span resolution. These budget workbooks are small (well under 1MB), so
    the memory/speed tradeoff of full-load mode is negligible here.
    """
    import openpyxl

    logger.info("Reading workbook %s", path)
    return openpyxl.load_workbook(path, data_only=True)


def segment_sheet_tables(sheet, gap_size: int = 1) -> List[Dict[str, int]]:
    """
    Split a sheet into table regions using a gap-based heuristic (ADR-0002,
    provisional/unvalidated): a region is a maximal run of rows that each
    have at least one non-empty cell, bounded by `gap_size` or more
    consecutive fully-empty rows or the sheet's end.
    """
    regions: List[Dict[str, int]] = []

    current_rows: List[int] = []
    current_min_col: Optional[int] = None
    current_max_col: Optional[int] = None
    empty_run = 0

    def close_region():
        if current_rows:
            regions.append({
                "min_row": current_rows[0],
                "max_row": current_rows[-1],
                "min_col": current_min_col,
                "max_col": current_max_col,
            })

    for row_idx, row in enumerate(sheet.iter_rows(), start=1):
        non_empty_cols = [cell.column for cell in row if cell.value is not None]

        if non_empty_cols:
            if empty_run >= gap_size:
                close_region()
                current_rows = []
                current_min_col = None
                current_max_col = None
            empty_run = 0
            current_rows.append(row_idx)
            row_min = min(non_empty_cols)
            row_max = max(non_empty_cols)
            current_min_col = row_min if current_min_col is None else min(current_min_col, row_min)
            current_max_col = row_max if current_max_col is None else max(current_max_col, row_max)
        else:
            empty_run += 1

    close_region()

    return regions


def extract_excel_tables(path: Path) -> Dict[str, Any]:
    """
    Extract segmented table regions with cell values/spans/header flags for
    every sheet in a workbook.

    Note: read_workbook() uses data_only=True, so a formula cell whose
    cached result was never saved by Excel (e.g. never opened/recalculated
    since editing) reads as None -> emitted here as "". This is a silent
    gap, not an error; budget workbooks are expected to hold pre-calculated
    values.

    Note: a merged cell whose top-left falls inside a region can extend
    past that region's max_col/max_row (segmentation is based on non-empty
    cells, not merge geometry) -- its emitted row_span/col_span may then
    exceed the region's declared bounds. Not corrected here per ADR-0002's
    "don't over-invest in heuristics" guidance; a real conflict would
    surface as a downstream normalize.py inconsistency to fix if it's ever
    hit on real data.
    """
    wb = read_workbook(path)

    sheets_out = []
    for sheet in wb.worksheets:
        regions = segment_sheet_tables(sheet)

        merged_ranges = list(sheet.merged_cells.ranges)

        regions_out = []
        for region in regions:
            rows_out = []
            for row_idx in range(region["min_row"], region["max_row"] + 1):
                row_cells = []
                is_header = row_idx == region["min_row"]
                for col_idx in range(region["min_col"], region["max_col"] + 1):
                    merged_range = None
                    for mr in merged_ranges:
                        if (mr.min_row <= row_idx <= mr.max_row
                                and mr.min_col <= col_idx <= mr.max_col):
                            merged_range = mr
                            break

                    if merged_range is not None:
                        is_top_left = (row_idx == merged_range.min_row
                                        and col_idx == merged_range.min_col)
                        if not is_top_left:
                            continue
                        row_span = merged_range.max_row - merged_range.min_row + 1
                        col_span = merged_range.max_col - merged_range.min_col + 1
                    else:
                        row_span = 1
                        col_span = 1

                    value = sheet.cell(row_idx, col_idx).value
                    row_cells.append({
                        "text": str(value) if value is not None else "",
                        "row_span": row_span,
                        "col_span": col_span,
                        "is_header": is_header,
                    })
                rows_out.append(row_cells)

            regions_out.append({
                "min_row": region["min_row"],
                "rows": rows_out,
            })

        sheets_out.append({
            "sheet_name": sheet.title,
            "regions": regions_out,
        })

    return {"sheets": sheets_out}
